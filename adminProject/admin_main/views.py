from django.shortcuts import render
from django.views.generic.list import ListView
from .models import Client_list
from .forms import ListForm
# Create your views here.
from django import forms
from django.views import generic
from django.http import HttpResponse
import pandas as pd
import os 
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import xlsxwriter
from openpyxl.writer.excel import save_virtual_workbook
from datetime import datetime
from django.utils.dateformat import DateFormat

class Admin_main(generic.TemplateView):
    def get(self, request, *args, **kwargs): #get 방식일 때 이렇게 움직여라!
        template_name='admin_main/index.html'
        return render(request, template_name)

class ClientList(generic.TemplateView):
    def get(self, request, *args, **kwargs):        
        template_name='admin_main/client_list.html'
        client_list = Client_list.objects.all()
        return render(request, template_name, {'client_list': client_list})

def check_post(request):
    template_name='admin_main/client_list_success.html'    
    if request.method=="POST":
        form = ListForm(request.POST)
        if form.is_valid() or not form.is_valid(): 
            list = form.save(commit=False)
            list.client_save()
            message="추가되었습니다."
            print(message)
            return render(request, template_name, {'message':message})
    else:
        template_name='admin_main/client_list_insert.html'
        form = ListForm
        return render(request,template_name,{"form":form})
        
#@csrf_exempt
def downolad_excel(request):
    if request.method=="POST":
        model = Client_list
        context_object_name ='admin_main'
        DBdata = list(model.objects.values())
        Datalist=[]

        wb = openpyxl.Workbook()
        sheet= wb.active
        num = 0
        sheet.cell(row=1, column=1).value = '인입시간'
        sheet.cell(row=1, column=2).value = '인입방법'
        sheet.cell(row=1, column=3).value = '회원번호'
        sheet.cell(row=1, column=4).value = '회원명'
        sheet.cell(row=1, column=5).value = '문서번호'
        sheet.cell(row=1, column=6).value = '제목'
        sheet.cell(row=1, column=7).value = '내용'  
        sheet.cell(row=1, column=8).value = '1차대응결과'
        sheet.cell(row=1, column=9).value = '1차대응담당자'
        sheet.cell(row=1, column=10).value = '장애보고서'
        sheet.cell(row=1, column=11).value = '파트너사'

        for list_data in DBdata:           
            Datalist.append(list_data)
            #배열을 통해 여러 열을 한번에 입력
            sheet.append([Datalist[num].get('voc_date'), Datalist[num].get('voc_method'), Datalist[num].get('client_number'), Datalist[num].get('client_name'), Datalist[num].get('voc_number'), Datalist[num].get('voc_title'), Datalist[num].get('voc_content'), Datalist[num].get('voc_comment'),Datalist[num].get('voc_manger'),Datalist[num].get('report'),Datalist[num].get('partner')])
            num+=1

        wb.save('test.xlsx')
        wb.close()
        filename = 'test'
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + filename + '.xlsx"'
        return response


class ClientList_detail(generic.DetailView):
    model = Client_list
    template_name='admin_main/client_list_detail.html'
    context_object_name = 'admin_main'

class ClientList_update(generic.UpdateView):
    model =Client_list
    fields =('voc_method','voc_date','client_number','client_name','voc_number','voc_title','voc_content','voc_comment','voc_manger','report','partner')
    template_name = 'admin_main/client_list_update.html'

    def form_valid(self, form):
        form.save()
        return render(self.request, 'admin_main/client_list_success.html',{"message":"수정되었습니다"})

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        print(self.object )
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        context = self.get_context_data(object= self.object, form =form)
        print(context)
        return self.render_to_response(context)

class ClientList_delete(generic.DeleteView):
    model = Client_list
    success_url = '/'
    context_object_name = 'admin_main'
